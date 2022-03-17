# -*- coding:utf-8 -*-
"""
    Excel文件读写
        获取excel文件
        获取sheet页
        获取单元格内容


    将所有关键字以字典的格式存放在列表中：
        key_li = [
            {"key":
                {"name":"name_value",
                "v":"v_value",
                "txt":"txt_value"}
                },
                {"name":"name_value",
                "v":"v_value",
                "txt":"txt_value"}
                },
                {"name":"name_value",
                "v":"v_value",
                "txt":"txt_value"}
            }
        ]

"""
import json
import openpyxl
from common.web_key import Key

# def read_excel():
#     # 获取excel文件
#     excel = openpyxl.load_workbook("../data/Excel/webui_data_demo1.xlsx")
#     # 获取所有的sheet页
#     sheets = excel.sheetnames
#     key_li = []
#     sheets_li = []
#     # 将操作步骤放入列表中，列表中每一项以字典格式保存数据（参数步骤数据），
#     dict_key = []
#     sh_key = dict.fromkeys(sheets_li)
#     for name in sheets:
#         sheets_li.append(name)
#         sheet = excel[name]
#         # print("*******{}*******".format(name))
#         # 遍历读取excel中的每个sheet页面内的数据
#         for values in sheet.values:
#             # print(values) # values:excel的每一行数据
#             if type(values[0]) is int:
#                 key_li.append(values[1])
#                 data_sheet = {
#                     "name": values[2],
#                     "value": values[3],
#                     "txt": values[4]
#                 }
#                 # 将参数中为None值全部去掉
#                 for keyvalue in list(data_sheet.keys()):
#                     if data_sheet[keyvalue] is None:
#                         del data_sheet[keyvalue]
#                 # print("data_sheet:", data_sheet)
#                 # 将所有关键字以字典的格式存放在列表中：
#                 li_c = {values[1]: data_sheet}
#                 dict_key.append(li_c)
#                 # print("li_c:", li_c)
#                 # print("dict_key:", dict_key)
#                 # print("————————————")
#         sh_key[name] = dict_key
#         # print("sh_key:", sh_key)
#         data = json.dumps(sh_key)
#         print("json数据格式data:", data)
#         return data
#
#
# # 具体执行excel测试用例
# def implement_case(data):
#     data = json.loads(data)
#     print(data)
#     for key in data:
#         # print(key)     # key:sheet页的名字
#         # print(data[key])        # data[key]:列表格式的用例执行步骤  --- [{步骤一数据},{步骤二数据}]   ---步骤数据的格式：{key:{参数}}
#         print("******正在执行WebUi:{}*******".format(key))
#         # 遍历
#         for sheet_key in data[key]:     # sheet_key:步骤数据:dict      cel_value:步骤数据的key
#             for cel_value in sheet_key:       # sheet_key[cel_value]:步骤参数:dict
#                 sheet_value = sheet_key[cel_value]
#                 if cel_value == "open_browser":
#                     # 当读取到open_browser时需要实例化
#                     key_value = Key(**sheet_value)
#                 else:
#                     print("key_value:00",key_value)
#                     print(type(key_value))
#                     print("sheet_value:00", sheet_value)
#                     print(type(sheet_value))
#                     getattr(key_value, cel_value)(**sheet_value)
#


# 获取excel文件
excel = openpyxl.load_workbook("../data/Excel/webui_data_demo1.xlsx")
# 获取sheet页信息
# sheet = excel['Sheet1']
# 获取指定单元格内容
# cell = sheet['A2'].value
# 获取所有的sheet页
sheets = excel.sheetnames
# 循环赋值sheet的名称
for name in sheets:
    sheet = excel[name]
    print("*******{}*******".format(name))
    # 获取整个sheet页的内容
    for values in sheet.values:
        # 进入测试用例正文
        if type(values[0]) is int:
            """
                提取内容：
                    操作行为
                    对应参数
            """
            data = {
                "name": values[2],
                "value": values[3],
                "txt": values[4]
            }
            # 将参数中为None值全部去掉
            for keyvalue in list(data.keys()):
                if data[keyvalue] is None:
                    del data[keyvalue]
            print(data)
            # 基于关键字行为进行测试内容的执行
            if values[1] == "open_browser":
                # 当读取到open_browser时需要实例化
                key = Key(**data)
            else:
                getattr(key, values[1])(**data)

# implement_case(read_excel())
