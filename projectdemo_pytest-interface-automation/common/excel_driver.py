# -*- coding:utf-8 -*-
"""
    Excel文件读写
        获取excel文件
        获取sheet页
        获取单元格内容
"""
import json
import openpyxl
from common.web_key import Key


# 读取excel的数据
def read_excel(path):
    # 获取excel文件
    excel = openpyxl.load_workbook(path)
    # 获取所有的sheet页的名称，列表的形式存放：sheets
    sheets = excel.sheetnames
    # 将所有的sheet页名称的列表转化为字典格式
    sh_key = dict.fromkeys(sheets)
    # 循环每个sheet页
    for name in sheets:
        # 获取所有sheet页的名称  sheet
        sheet = excel[name]
        # 将每个sheet页的操作步骤放入列表中，列表中每一项以字典格式保存数据（参数步骤数据），
        dict_key = []
        # print("*******{}*******".format(name))
        # 遍历读取excel中的每个sheet页面内的数据
        for values in sheet.values:  # values:excel的每一行数据
            # print(values)
            if type(values[0]) is int:
                # 将所有关键字添加到列表 key_li 中
                # key_li.append(values[1])
                data_sheet = {
                    "name": values[2],
                    "value": values[3],
                    "txt": values[4]
                }
                # 将参数中为None值全部去掉
                for keyvalue in list(data_sheet.keys()):
                    if data_sheet[keyvalue] is None:
                        del data_sheet[keyvalue]
                # 将所有关键字以及对应的参数数据,步骤描述   以字典（li_c）的格式存放在dict_key列表中：
                li_c = {"describe": values[5], values[1]: data_sheet}
                dict_key.append(li_c)
                # 将每个sheet页的数据列表赋值到sheet页字典的value中
                sh_key[name] = dict_key
    data = json.dumps(sh_key)
    print("json数据格式data:", data)
    return data


# 具体执行excel测试用例
def implement_case(data):
    data = json.loads(data)
    print(data)
    # 遍历每个sheet页面数据
    for key in data:
        # print(key)     # key:sheet页的名字
        # print(data[key])        # data[key]:列表格式的用例执行步骤  --- [{步骤一数据},{步骤二数据}]   ---步骤数据的格式：{key:{参数},describe:描述}
        print("")
        print("******正在执行WebUi:{}*******".format(key))
        # 遍历每个步骤
        for sheet_key in data[key]:  # sheet_key:步骤数据:dict      cel_value:步骤数据的key
            # 将步骤描述赋值给describe
            describe = sheet_key["describe"]
            del sheet_key["describe"]
            # 遍历每个步骤参数字典里的key
            for cel_value in sheet_key:  # sheet_key[cel_value]:步骤参数:dict
                sheet_value = sheet_key[cel_value]
                if cel_value == "open_browser":
                    # 当读取到open_browser时需要实例化
                    key_value = Key(**sheet_value)
                else:
                    try:
                        getattr(key_value, cel_value)(**sheet_value)
                        print("Message:执行 {}--SUCCESS".format(describe))
                    except Exception as e:
                        print("Message:执行 {}>>>>FAIL".format(describe))
                        print(e)


'''
# 获取excel文件
excel = openpyxl.load_workbook("../data/Excel/webui_data_demo1.xlsx")
# 获取sheet页信息
# sheet = excel['Sheet1']
# 获取指定单元格内容
# cell = sheet['A2'].value
# 获取所有的sheet页
sheets = excel.sheetnames
# 循环赋值sheet的名称
for name in sheets:
    sheet = excel[name]
    print("*******{}*******".format(name))
    # 获取整个sheet页的内容
    for values in sheet.values:
        # 进入测试用例正文
        if type(values[0]) is int:
            """
                提取内容：
                    操作行为
                    对应参数
            """
            data = {
                "name": values[2],
                "value": values[3],
                "txt": values[4]
                }
            # 将参数中为None值全部去掉
            for keyvalue in list(data.keys()):
                if data[keyvalue] is None:
                    del data[keyvalue]
            print(data)
            # 基于关键字行为进行测试内容的执行
            if values[1] == "open_browser":
                # 当读取到open_browser时需要实例化
                key = Key(**data)
            else:
                getattr(key, values[1])(**data)


'''
